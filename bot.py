import openpyxl as opx
import telebot

# Токен бота
token = ""

# Функция самого бота
def telegram_bot(token):
    bot = telebot.TeleBot(token)

    # Функция стартового сообщение от бота
    @bot.message_handler(commands=["start"])
    def start_message(message):
        bot.send_message(
            message.chat.id, "👋 Привет, если хочешь узнать свою оценку, напиши мне свое ФИО и группу. В будущем ты сможешь просто писать 'Оценки'")
    
    # Функция приема сообщений пользователя и ответа на них
    @bot.message_handler(content_types=["text"])
    def send_message(message):
        #Обработка сообщения "Оценки"
        if message.text.lower() == "оценки":
            try:
                excel = opx.open("оценки.xlsx")
                m = None

                #Цикл по перебору таблиц
                for row1 in excel.sheetnames:
                    t_name = row1
                    tables = excel[t_name.upper()]

                    #Перебор значений в таблице, и последующая их запись в переменные для вывода пользователю
                    for names in range(1, tables.max_row + 1):
                        marked = tables[names][1].value
                        u_id = message.from_user.id
                        scored = tables[names][2].value
                        print(scored)
                        if u_id == scored:
                            m = marked
                
                #Проверка переменной с оценкой, которая выдает "None" в случае если айдишник пользователя не был найден
                if m == None:
                    bot.send_message(message.chat.id, f"Обратитесь к администратору")
                else:
                    bot.send_message(message.chat.id, f"Твоя оценка на данный момент - {m} 👍")

            except IndexError:
                pass
        else:
            try:
                # Дробление полученного сообщения, для приминения поиска нужные параметров в файле
                chat = message.text.split()
                text = chat[0]
                text1 = chat[1]
                text2 = chat[2]
                name = f"{text} {text1} {text2}"
                group = chat[3]

                # Открытие xlsx файла и поиск нужной таблицы
                ex = opx.open("оценки.xlsx")
                table = ex[group.upper()]
                #Запись айдишника пользователя
                user_id = message.from_user.id            

                # Цикл по перебору значений в таблице
                for row in range(1, table.max_row + 1):
                    # Переменные куда записываеться ФИО на данной итерации для будущего сравнения и оценка, для её записывания в перменную, в случае совпадения ФИО
                    FIO = table[row][0].value
                    score = table[row][1].value

                    if name.upper() in FIO.upper() or FIO.upper() == name.upper():
                        mark = score
                        table.cell(row = row, column = 3).value = user_id
                        ex.save("Оценки.xlsx")

                bot.send_message(
                    message.chat.id, f"Твоя оценка на данный момент - {mark} 👍")

            # Ловля ошибок, с которыми может столкнуться пользователь
            except KeyError:
                bot.send_message(message.chat.id, f"Эм 😒 посмотри, ты похоже ошибся в группе, возможно забыл поставить знак '-' ")
            except UnboundLocalError:
                bot.send_message(message.chat.id, f"Там это 😅 такого человека не существует, проверька, не ошибся-ли ты часом ")
            except IndexError:
                bot.send_message(message.chat.id, f"Я ПРОСИЛ, ФАМИЛИЮ, ИМЯ, ОТЧЕСТВО И ГРУППУ 👿 ")

    bot.polling()


telegram_bot(token)
